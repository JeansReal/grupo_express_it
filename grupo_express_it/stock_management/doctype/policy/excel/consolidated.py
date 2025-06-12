from io import BytesIO

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import frappe
from frappe.desk.utils import provide_binary_file
from grupo_express_it.stock_management.doctype.policy.policy import Policy


def _policy(ws: Worksheet, policy: Policy) -> None:
	ws.cell(row=1, column=1, value=policy.name)

	for (cell, label) in [
		("A1", "Registro:"),       ("B2", policy.name),
		("A2", "Proveedor:"),      ("B3", policy.provider),
		("A4", "Fecha:"),          ("B4", str(policy.posting_date)),
		("I2", "Póliza:"),         ("J2", policy.policy),
		("I3", "Factura:"),        ("J3", policy.invoice),
		("I4", "Tasa de Cambio:"), ("J4", policy.exchange_rate)
	]:
		ws[cell] = label
	pass


@frappe.whitelist(allow_guest=False)
def download(policies: str):

	wb = Workbook()
	ws = wb.active          # Get Current Sheet

	for policy in policies.split(','):
		doc = Policy('Policy', policy)

		_policy(ws=ws, policy=doc)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	provide_binary_file('Consolidado', 'xlsx', xlsx_file.getvalue())  # Build xlsx Response File -> See xlsxutils.py
