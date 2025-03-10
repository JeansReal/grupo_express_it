from io import BytesIO

from google.api_core.datetime_helpers import from_iso8601_date
from openpyxl.styles import Alignment, Font, NamedStyle, Side, Border, PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import frappe
from frappe.desk.utils import provide_binary_file
from grupo_express_it.stock_management.doctype.policy.policy import Policy
from grupo_express_it.stock_management.doctype.policy_cif_cost.policy_cif_cost import PolicyCIFCost
from grupo_express_it.stock_management.doctype.policy_item.policy_item import PolicyItem
from grupo_express_it.stock_management.doctype.policy_nationalization_cost.policy_nationalization_cost import \
	PolicyNationalizationCost

# Default Columns for Policy Child Tables: CIF and Nationalization Costs
CHILD_COLUMNS = [
	{'col': 'B',                                                                             'label': 'PROVEEDOR', 'attr': 'provider'},
	{'col': 'C', 'merge': 'D',                                                               'label': 'DESCRIPCIÓN', 'attr': 'description'},
	{'col': 'E', 'style': 'date',                                                            'label': 'FECHA', 'attr': 'posting_date'},
	{'col': 'G', 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'a9d18e', 'label': 'MONTO $', 'attr': 'amount_usd'},
	{'col': 'H', 'style': 'currency',                                                        'label': 'TCO', 'attr': 'exchange_rate'},
]


def _register_styles(wb: Workbook) -> None:
	number_format = '#,##0.00'
	font_bold = Font(bold=True, size=10)
	side = Side(style='medium')                                    # Default Side for the Border
	border = Border(left=side, right=side, top=side, bottom=side)  # Default Border

	table_header_style = {
		'font': font_bold, 'border': border,
		'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
	}
	number_style = {
		'font': Font(size=10), 'number_format': number_format, 'alignment': Alignment(horizontal='right', indent=1)
	}

	header = NamedStyle(name='table-header', **table_header_style)
	pre_header = NamedStyle(name='table-pre-header', **table_header_style, fill=PatternFill('solid', fgColor='bdd7ee'))

	text_center = NamedStyle(name='text-center', alignment=Alignment(horizontal='center'))
	date = NamedStyle(name='date', number_format='DD/MM/YYYY;@', alignment=Alignment(horizontal='center'))

	number = NamedStyle(name='number', **number_style)
	currency = NamedStyle(name='currency', **{**number_style, 'number_format': '#,##0.00##'})  # Adding 4 Decimal Places
	total_style = NamedStyle(name='table-total', **{**number_style, 'font': font_bold}, border=border)

	[wb.add_named_style(style=style) for style in [header, pre_header, text_center, date, number, currency, total_style]]


def _table_pre_header(ws: Worksheet, row: int, col: int, value: str, col_range: [str, str]) -> int:
	ws.cell(row=row, column=col, value=value).style = 'table-pre-header'
	ws.merge_cells(f'{col_range[0]}{row}:{col_range[1]}{row}')
	return row + 1


def _table_builder(ws: Worksheet, current_row: int, columns: list[dict], data: list[any]) -> int:
	data_start = current_row + 1  # Data starts a row after the Header

	for col in columns:  # Table Header
		ws.cell(row=current_row, column=column_index_from_string(col['col']), value=col['label']).style = 'table-header'

		if col_merge := col.get('merge', False):
			ws.merge_cells(f'{col['col']}{current_row}:{col_merge}{current_row}')

	for item in data:  # Table Content(Body)
		current_row += 1  # Adding a row to avoid overwriting the header

		for col in columns:
			if col['attr'][0] == '=':  # If it's a formula
				value = col['attr'].format(row=current_row)
			else:
				value = getattr(item, col['attr'], "")

			cell = ws.cell(row=current_row, column=column_index_from_string(col['col']), value=value)  # Cell Value
			cell.style = col.get('style', 'Normal')  # Cell Style

			if col_merge := col.get('merge', False):
				ws.merge_cells(f'{col['col']}{current_row}:{col_merge}{current_row}')

			if col_fill := col.get('fill'):
				cell.fill = PatternFill("solid", fgColor=col_fill)
				cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

	current_row += 1  # We add a row for the Total
	for col in columns:  # Total Row
		if col.get('total'):
			cell_value = f'=SUM({col["col"]}{data_start}:{col["col"]}{current_row - 1})'
			cell = ws.cell(row=current_row, column=column_index_from_string(col["col"]), value=cell_value)
			cell.style = 'table-total'

			if total_fill := col.get('total_fill'):
				cell.fill = PatternFill("solid", fgColor=total_fill)

	ws.cell(row=current_row, column=2, value="TOTAL").style = 'table-header'
	ws.merge_cells(f'B{current_row}:D{current_row}')

	return current_row  # Returns the Total Row


def _policy_header(ws: Worksheet, policy: Policy) -> None:
	for (cell, label) in [
		("A1", "Liquidación de Importación"),
		("A2", "Registro:"),       ("B2", policy.name),
		("A3", "Proveedor:"),      ("B3", policy.provider),
		("A4", "Fecha:"),          ("B4", str(policy.posting_date)),
		("I2", "Póliza:"),         ("J2", policy.policy),
		("I3", "Factura:"),        ("J3", policy.invoice),
		("I4", "Tasa de Cambio:"), ("J4", policy.exchange_rate)
	]:
		ws[cell] = label

	ws.merge_cells('A1:N1')
	ws['A1'].style = "Title"
	ws['A1'].alignment = Alignment(horizontal='center')

	ws['I4'].alignment = Alignment(horizontal='right')
	ws['J4'].number_format = 'C$ 00.000000'  # Currency Exchange Format


def _policy_items(ws: Worksheet, current_row: int, policy_items: list[PolicyItem]) -> int:
	columns = [
		{'col': 'A', 'width': 10, 'style': 'number',                   'total': True,                         'label': 'CANT.', 'attr': 'qty'},
		{'col': 'B', 'width': 60,                                                                             'label': 'DESCRIPCIÓN', 'attr': 'item'},
		{'col': 'C', 'width': 12, 'style': 'text-center',                                                     'label': 'U.M.', 'attr': 'uom'},
		{'col': 'D', 'width': 14, 'style': 'number',                                                          'label': 'COSTO UNITARIO FOB $', 'attr': 'fob_unit_price'},
		{'col': 'E', 'width': 12, 'style': 'number',                   'total': True,                         'label': 'FOB DÓLARES $', 'attr': '=A{row}*D{row}'},
		{'col': 'F', 'width': 12, 'style': 'number',                   'total': True,                         'label': 'FLETES $', 'attr': 'freight_cost'},
		{'col': 'G', 'width': 10, 'style': 'number',                   'total': True,                         'label': 'SEGUROS $', 'attr': 'insurance_cost'},
		{'col': 'H', 'width': 11, 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'a9d18e', 'label': 'CIF DÓLARES $', 'attr': '=E{row}+F{row}+G{row}'},
		{'col': 'I', 'width': 12, 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'f2f2f2', 'label': 'CIF CORDOBAS C$', 'attr': '=H{row}*$J$4'},
		{'col': 'J', 'width': 12, 'style': 'number',                   'total': True,                         'label': 'IMPUESTOS ADUANEROS C$', 'attr': 'customs_taxes'},
		{'col': 'K', 'width': 12, 'style': 'number',                   'total': True,                         'label': 'COSTOS DE NAC. C$', 'attr': 'nationalization_total'},
		{'col': 'L', 'width': 15, 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'f2f2f2', 'label': 'TOTAL COSTOS NACIONAL. C$', 'attr': '=J{row}+K{row}'},
		{'col': 'M', 'width': 15, 'style': 'number', 'fill': 'deebf7', 'total': True, 'total_fill': 'a9d18e', 'label': 'COSTO TOTAL PRODUCTO C$', 'attr': '=I{row}+L{row}'},
		{'col': 'N', 'width': 12, 'style': 'number', 'fill': 'deebf7',                                        'label': 'COSTO UNITARIO C$', 'attr': '=M{row}/A{row}'}
	]

	# Custom Alignments
	[setattr(ws.column_dimensions[col], 'alignment', Alignment(horizontal='center')) for col in ['E', 'F']]

	ws.row_dimensions[current_row].height = 40  # Set Header Row Height
	for col in columns:  # Set Column Widths Relative to Policy Items
		ws.column_dimensions[col['col']].width = col['width']

	return _table_builder(ws=ws, current_row=current_row, columns=columns, data=policy_items)


def _policy_cif_costs(ws: Worksheet, current_row: int, cif_costs: list[PolicyCIFCost]) -> int:
	columns = CHILD_COLUMNS + [
		{'col': 'F', 'label': 'REFERENCIA', 'attr': 'type'},
		{'col': 'I', 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'bdd7ee', 'label': 'TOTAL C$', 'attr': '=G{row}*H{row}'}
	]

	current_row = _table_pre_header(ws=ws, row=current_row, col=2, value='VALOR CIF (FOB + FLETES + SEGUROS)', col_range=['B', 'I'])
	return _table_builder(ws=ws, current_row=current_row, columns=columns, data=cif_costs)


def _policy_nationalization_cost(ws: Worksheet, current_row: int, nationalization_cost: list[PolicyNationalizationCost]) -> int:
	columns = CHILD_COLUMNS + [
		{'col': 'F',                                                                             'label': 'REFERENCIA', 'attr': 'reference'},
		{'col': 'I', 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'bdd7ee', 'label': 'SUB TOTAL', 'attr': 'amount_nio'},  # FIXME: this sould be a formula, but sometimes the amount_usd(col[G]) is not set
		{'col': 'J', 'style': 'number',                   'total': True, 'total_fill': 'f2f2f2', 'label': 'I.V.A', 'attr': 'iva'},
		{'col': 'K', 'style': 'number',                   'total': True, 'total_fill': 'f2f2f2', 'label': 'RETENCION', 'attr': '=0'},  # Virtual
		{'col': 'L', 'style': 'number', 'fill': 'f2f2f2', 'total': True, 'total_fill': 'bdd7ee', 'label': 'MONTO A PAGAR', 'attr': '=SUM(I{row}:K{row})'}
	]

	current_row = _table_pre_header(ws=ws, row=current_row, col=2, value='GASTOS DE NACIONALIZACION', col_range=['B', 'L'])
	return _table_builder(ws=ws, current_row=current_row, columns=columns, data=nationalization_cost)


def _total(ws: Worksheet, current_row: int, cif_total_row: int, nat_total_row: int):
	# This is a little Hacky, but it works: It must be formulas
	current_row = _table_pre_header(ws=ws, row=current_row, col=2, value='TOTAL COSTO DE MERCADERIA + COSTOS NACIONALIZACION', col_range=['B', 'F']) - 1  # pre-header returns next row

	ws.cell(row=current_row, column=column_index_from_string('G'), value=f'=G{cif_total_row} + G{nat_total_row}').style = 'table-total'
	ws.cell(row=current_row, column=column_index_from_string('I'), value=f'=I{cif_total_row} + I{nat_total_row}').style = 'table-total'
	ws.cell(row=current_row, column=column_index_from_string('J'), value=f'=J{nat_total_row}').style = 'table-total'
	ws.cell(row=current_row, column=column_index_from_string('L'), value=f'=I{cif_total_row} + L{nat_total_row}').style = 'table-total'


@frappe.whitelist(allow_guest=False)
def download(policy: str):
	policy = Policy('Policy', policy)

	wb = Workbook()
	ws = wb.active          # Get Current Sheet
	ws.title = policy.name  # Rename Sheet

	_register_styles(wb)
	_policy_header(ws, policy)

	# Adding Extra Row for it to make sense
	policy.append(key='cif_costs', value={
		'provider': policy.provider,
		'description': 'Mercaderia variada',
		'posting_date': policy.posting_date,
		'type': policy.invoice,  # Field reference: Does not exist on PolicyCIFCost
		'amount_usd': policy.total_fob,
		'exchange_rate': policy.exchange_rate  # This
	}, position=0)

	current_row = _policy_items(ws, current_row=8, policy_items=policy.items)
	cif_total_row = _policy_cif_costs(ws, current_row=current_row + 4, cif_costs=policy.cif_costs)
	nat_total_row = _policy_nationalization_cost(ws, current_row=cif_total_row + 4, nationalization_cost=policy.nationalization_costs)
	_total(ws, current_row=nat_total_row + 3, cif_total_row=cif_total_row, nat_total_row=nat_total_row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	provide_binary_file(policy.name, 'xlsx', xlsx_file.getvalue())  # Build xlsx Response File -> See xlsxutils.py
